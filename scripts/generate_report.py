#!/usr/bin/env python3
"""Generate YonClaw skill acceptance Excel report from staged artifacts.

Reads skill-info.json, checkpoints-covered.json and test-results.json,
then produces a .xlsx report conforming to the yonclaw-skill-acceptance
Excel format specification (4 sheets, TARGET/READ/VERDICT exclusion, etc.).

Usage:
    python3 generate_report.py \\
        --skill-info <path>/skill-info.json \\
        --checkpoints <path>/checkpoints-covered.json \\
        --test-results <path>/test-results.json \\
        [--output <path>/report.xlsx] \\
        [--template <path>/checkpoints.json]
"""

import json
import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

CHECKPOINT_ONLY_CATEGORIES = {"目标确认", "强制读取", "结论门禁", "报告交付自检"}

"""
Ordered display categories for skill-summary / detail sheets.
Names MUST match checkpoints.json category values exactly.
Categories in CHECKPOINT_ONLY_CATEGORIES are excluded from detail/summary.
"""
DISPLAY_ORDER = [
    "BIP 规范专项检查",
    "通用规范检查",
    "安全与风险审计",
    "异常输入与副作用审计",
    "平台集成检查",
    "功能清单与覆盖",
    "动态用例结果",
    "BIP 脚本合规检查",
    "Baseline Failure",
]

VALID_DETAIL_RESULTS = {"通过", "不通过", "未执行"}
VALID_CHECK_RESULTS = {"通过", "未通过", "阻塞", "不适用"}

STATUS_TO_RESULT = {
    "covered": "通过",
    "failed": "不通过",
    "blocked": "未执行",
    "skipped": "未执行",
    "not_applicable": "不通过",
}

STATUS_TO_SELF_CHECK = {
    "covered": "通过",
    "failed": "未通过",
    "blocked": "阻塞",
    "skipped": "未通过",
    "not_applicable": "不适用",
}

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
THIN_BORDER = Border(
    top=Side(style="thin"),
    left=Side(style="thin"),
    bottom=Side(style="thin"),
    right=Side(style="thin"),
)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")

# ---------------------------------------------------------------------------
# JSON loading helpers
# ---------------------------------------------------------------------------


def load_json(path):
    if not os.path.isfile(path):
        print(f"ERROR: file not found: {path}", file=sys.stderr)
        sys.exit(1)
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as exc:
        print(f"ERROR: invalid JSON {path}: {exc}", file=sys.stderr)
        sys.exit(1)


def as_list(value):
    return value if isinstance(value, list) else []


# ---------------------------------------------------------------------------
# Data transformation
# ---------------------------------------------------------------------------


def category_for_checkpoint(cp, template_map):
    cp_id = str(cp.get("id", ""))
    if cp_id and cp_id in template_map:
        return template_map[cp_id].get("category", "")
    return cp.get("category", "")


def detail_category(raw_category):
    """Use the original category name from checkpoints.json as-is.

    CHECKPOINT_ONLY_CATEGORIES are filtered out before reaching detail rows,
    so they won't appear in detail/summary sheets regardless.
    """
    return raw_category


def build_detail_rows(checkpoints, template_map):
    rows = []
    for cp in checkpoints:
        raw_cat = category_for_checkpoint(cp, template_map)
        display_cat = detail_category(raw_cat)

        if raw_cat in CHECKPOINT_ONLY_CATEGORIES:
            continue

        status = str(cp.get("status", "")).strip()
        cp_id = str(cp.get("id", ""))
        checkpoint_text = cp.get("checkpoint") or (template_map.get(cp_id, {}).get("checkpoint", ""))

        if status == "not_applicable":
            test_result = "未执行"
        elif status in STATUS_TO_RESULT:
            test_result = STATUS_TO_RESULT[status]
        else:
            test_result = "未执行"

        note_parts = []
        if cp.get("evidence"):
            note_parts.append(str(cp["evidence"]))
        if cp.get("reason"):
            note_parts.append(str(cp["reason"]))
        note = "; ".join(note_parts)

        rows.append({
            "方法论大项": display_cat,
            "具体测试点": str(checkpoint_text),
            "测试结果": test_result,
            "备注": note,
        })
    return rows


# ---------------------------------------------------------------------------
# Excel generation
# ---------------------------------------------------------------------------


def apply_header_style(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN


def apply_cell_border(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = THIN_BORDER
        cell.alignment = WRAP_ALIGN


def auto_width(ws, col_count, max_width=60):
    for col in range(1, col_count + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                val = str(cell.value or "")
                max_len = max(max_len, min(len(val), max_width))
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = max(max_len + 4, 12)


def generate_report(skill_info, checkpoints_data, test_results, template_data, output_path):
    wb = Workbook()

    skill_name = (skill_info.get("skill") or {}).get("name", "unknown-skill")
    checkpoints = as_list(checkpoints_data.get("checkpoints"))
    template_map = {}
    if template_data:
        for item in as_list(template_data.get("checkpoints")):
            cp_id = str(item.get("id", ""))
            if cp_id:
                template_map[cp_id] = item

    detail_sheet_name = skill_name

    # ------------------------------------------------------------------
    # Build detail rows first to get statistics
    # ------------------------------------------------------------------
    detail_rows = build_detail_rows(checkpoints, template_map)

    stats = {"通过": 0, "不通过": 0, "未执行": 0}
    for row in detail_rows:
        if row["测试结果"] in VALID_DETAIL_RESULTS:
            stats[row["测试结果"]] += 1

    total_count = stats["通过"] + stats["不通过"] + stats["未执行"]
    pass_rate = f"{(stats['通过'] / total_count * 100):.1f}%" if total_count > 0 else "0%"

    # ------------------------------------------------------------------
    # Sheet 1: 汇总
    # ------------------------------------------------------------------
    # Remove default sheet
    wb.remove(wb.active)

    ws_summary = wb.create_sheet("汇总")
    summary_headers = ["技能清单", "测试点总数", "验证通过", "验证不通过", "未执行", "通过率"]
    for col, header in enumerate(summary_headers, 1):
        ws_summary.cell(row=1, column=col, value=header)
    apply_header_style(ws_summary, 1, len(summary_headers))

    ws_summary.cell(row=2, column=1, value=skill_name)
    ws_summary.cell(row=2, column=2, value=total_count)
    ws_summary.cell(row=2, column=3, value=stats["通过"])
    ws_summary.cell(row=2, column=4, value=stats["不通过"])
    ws_summary.cell(row=2, column=5, value=stats["未执行"])
    ws_summary.cell(row=2, column=6, value=pass_rate)
    apply_cell_border(ws_summary, 2, len(summary_headers))
    auto_width(ws_summary, len(summary_headers))

    # ------------------------------------------------------------------
    # Sheet 2: <skill-name>汇总
    # ------------------------------------------------------------------
    ws_skill_summary = wb.create_sheet(f"{skill_name}汇总")
    skill_summary_headers = ["项目", "测试点总数", "验证通过", "验证不通过", "未执行", "通过率"]
    for col, header in enumerate(skill_summary_headers, 1):
        ws_skill_summary.cell(row=1, column=col, value=header)
    apply_header_style(ws_skill_summary, 1, len(skill_summary_headers))

    by_category = {}
    for row in detail_rows:
        cat = row["方法论大项"]
        if cat not in by_category:
            by_category[cat] = {"通过": 0, "不通过": 0, "未执行": 0}
        if row["测试结果"] in VALID_DETAIL_RESULTS:
            by_category[cat][row["测试结果"]] += 1

    skill_summary_row = 2
    for display_cat in DISPLAY_ORDER:
        cat_stats = by_category.get(display_cat)
        if not cat_stats:
            continue
        cat_total = cat_stats["通过"] + cat_stats["不通过"] + cat_stats["未执行"]
        cat_rate = f"{(cat_stats['通过'] / cat_total * 100):.1f}%" if cat_total > 0 else "0%"

        ws_skill_summary.cell(row=skill_summary_row, column=1, value=display_cat)
        ws_skill_summary.cell(row=skill_summary_row, column=2, value=cat_total)
        ws_skill_summary.cell(row=skill_summary_row, column=3, value=cat_stats["通过"])
        ws_skill_summary.cell(row=skill_summary_row, column=4, value=cat_stats["不通过"])
        ws_skill_summary.cell(row=skill_summary_row, column=5, value=cat_stats["未执行"])
        ws_skill_summary.cell(row=skill_summary_row, column=6, value=cat_rate)
        apply_cell_border(ws_skill_summary, skill_summary_row, len(skill_summary_headers))
        skill_summary_row += 1
    auto_width(ws_skill_summary, len(skill_summary_headers))

    # ------------------------------------------------------------------
    # Sheet 3: <skill-name> (detail)
    # ------------------------------------------------------------------
    ws_detail = wb.create_sheet(detail_sheet_name)
    detail_headers = ["方法论大项", "具体测试点", "测试结果", "备注"]
    for col, header in enumerate(detail_headers, 1):
        ws_detail.cell(row=1, column=col, value=header)
    apply_header_style(ws_detail, 1, len(detail_headers))

    detail_row_offset = 2
    for i, row in enumerate(detail_rows):
        row_num = detail_row_offset + i
        ws_detail.cell(row=row_num, column=1, value=row["方法论大项"])
        ws_detail.cell(row=row_num, column=2, value=row["具体测试点"])
        ws_detail.cell(row=row_num, column=3, value=row["测试结果"])
        ws_detail.cell(row=row_num, column=4, value=row["备注"])
        apply_cell_border(ws_detail, row_num, len(detail_headers))
    auto_width(ws_detail, len(detail_headers))
    ws_detail.column_dimensions["B"].width = 50
    ws_detail.column_dimensions["D"].width = 60

    # ------------------------------------------------------------------
    # Sheet 4: 检查点自检
    # ------------------------------------------------------------------
    ws_check = wb.create_sheet("检查点自检")
    check_headers = ["检查点ID", "方法论大项", "检查点", "适用性", "覆盖位置", "自检结果", "备注"]
    for col, header in enumerate(check_headers, 1):
        ws_check.cell(row=1, column=col, value=header)
    apply_header_style(ws_check, 1, len(check_headers))

    # Build detail row map: checkpoint id -> detail row number
    detail_row_map = {}
    current_row = detail_row_offset
    for cp in checkpoints:
        raw_cat = category_for_checkpoint(cp, template_map)
        if raw_cat in CHECKPOINT_ONLY_CATEGORIES:
            continue
        cp_id = str(cp.get("id", ""))
        detail_row_map[cp_id] = current_row
        current_row += 1

    # Build self-check rows
    check_row = 2
    for cp in checkpoints:
        cp_id = str(cp.get("id", ""))
        raw_cat = category_for_checkpoint(cp, template_map)
        display_cat = detail_category(raw_cat)
        checkpoint_text = cp.get("checkpoint") or template_map.get(cp_id, {}).get("checkpoint", "")
        status = str(cp.get("status", "")).strip()

        self_check = STATUS_TO_SELF_CHECK.get(status, "阻塞")
        applicability = "不适用" if status == "not_applicable" else "适用"

        if raw_cat in CHECKPOINT_ONLY_CATEGORIES:
            coverage_location = "检查点自检" if self_check == "通过" else "未覆盖"
        elif cp_id in detail_row_map:
            row_num = detail_row_map[cp_id]
            coverage_location = f"{detail_sheet_name}!A{row_num}"
        else:
            coverage_location = "未覆盖"

        note_parts = []
        if cp.get("evidence"):
            note_parts.append(str(cp["evidence"]))
        if cp.get("reason"):
            note_parts.append(str(cp["reason"]))
        note = "; ".join(note_parts)

        ws_check.cell(row=check_row, column=1, value=cp_id)
        ws_check.cell(row=check_row, column=2, value=display_cat)
        ws_check.cell(row=check_row, column=3, value=checkpoint_text)
        ws_check.cell(row=check_row, column=4, value=applicability)
        ws_check.cell(row=check_row, column=5, value=coverage_location)
        ws_check.cell(row=check_row, column=6, value=self_check)
        ws_check.cell(row=check_row, column=7, value=note)
        apply_cell_border(ws_check, check_row, len(check_headers))
        check_row += 1
    auto_width(ws_check, len(check_headers))
    ws_check.column_dimensions["C"].width = 50
    ws_check.column_dimensions["G"].width = 60

    # ------------------------------------------------------------------
    # Save
    # ------------------------------------------------------------------
    out_dir = os.path.dirname(output_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    wb.save(output_path)
    print(f"Report generated: {output_path}")
    return output_path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def parse_args(argv):
    args = {"template": None, "output": None}
    i = 1
    while i < len(argv):
        if argv[i] == "--skill-info" and i + 1 < len(argv):
            args["skill_info"] = argv[i + 1]; i += 2
        elif argv[i] == "--checkpoints" and i + 1 < len(argv):
            args["checkpoints"] = argv[i + 1]; i += 2
        elif argv[i] == "--test-results" and i + 1 < len(argv):
            args["test_results"] = argv[i + 1]; i += 2
        elif argv[i] == "--template" and i + 1 < len(argv):
            args["template"] = argv[i + 1]; i += 2
        elif argv[i] in ("--output", "-o") and i + 1 < len(argv):
            args["output"] = argv[i + 1]; i += 2
        else:
            i += 1
    return args


def main():
    args = parse_args(sys.argv)

    if not args.get("skill_info") or not args.get("checkpoints") or not args.get("test_results"):
        print("Usage: python3 generate_report.py --skill-info <path> --checkpoints <path> --test-results <path> [--output <path>] [--template <path>]", file=sys.stderr)
        sys.exit(1)

    skill_info = load_json(args["skill_info"])
    checkpoints_data = load_json(args["checkpoints"])
    test_results = load_json(args["test_results"])

    default_template = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "assets", "checkpoints.json")
    template_path = args.get("template") or default_template
    template_data = load_json(template_path) if os.path.isfile(template_path) else None

    skill_name = (skill_info.get("skill") or {}).get("name", "unknown-skill")
    output_path = args.get("output") or f"{skill_name}-acceptance-report.xlsx"

    generate_report(skill_info, checkpoints_data, test_results, template_data, output_path)


if __name__ == "__main__":
    main()
